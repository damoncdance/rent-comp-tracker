"""Export snapshots to Excel workbooks.

Two export modes:
  1. Per-property: data/exports/<slug>_Availability.xlsx (3 sheets)
  2. Comp report:  data/exports/Rent_Comp_Report.xlsx — all properties in one
     workbook with a Comp Overview coversheet + per-property detail sheets.

Uses openpyxl. Called from daily_run after snapshots complete.
"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from src.config import DATA_DIR, PROPERTY_NAME, PROPERTY_URL, get_property_by_id
from src.fees import FEES
from src.storage import db, latest_snapshot_id


# --- Styles ------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFFFF")
_TITLE_FONT  = Font(bold=True, size=16)
_SUB_FONT    = Font(size=11, color="FF5C5C5C")
_SECTION_FONT = Font(bold=True, size=11)
_SUBJECT_FILL = PatternFill(start_color="FFE8E0FF", end_color="FFE8E0FF", fill_type="solid")
_SUBJECT_FONT = Font(bold=True, size=11, color="FF4B3D99")
_AVG_FILL = PatternFill(start_color="FFF0F0F0", end_color="FFF0F0F0", fill_type="solid")
_AVG_FONT = Font(bold=True, size=11)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="FFD0D0D0"),
)
_DATE_FMT    = 'mmm d, yyyy'
_MONEY_FMT   = '$#,##0'
_NUM_FMT     = '#,##0'
_PSF_FMT     = '$#,##0.00'
_PCT_FMT     = '0.0%'

EXPORTS_DIR = DATA_DIR / "exports"


def report_workbook_path(prop: dict | None = None) -> Path:
    """Path where the latest availability workbook is written."""
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    name = prop["name"] if prop else PROPERTY_NAME
    slug = _slugify(name)
    return EXPORTS_DIR / f"{slug}_Availability.xlsx"


def _bed_label(beds: int) -> str:
    if beds == 0:
        return "Studio"
    return f"{beds} Bedroom"


def _slugify(name: str) -> str:
    import re
    slug = re.sub(r'[^a-zA-Z0-9\s-]', '', name)
    return slug.strip().replace(" ", "_")


def _apply_header_row(ws, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, min_width: float = 8.0) -> None:
    """Set column widths based on content, with reasonable overrides."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, 35)


# --- Public API --------------------------------------------------------------

def export(snapshot_id: int | None = None, prop: dict | None = None) -> str | None:
    """Build the xlsx from a snapshot. Returns the output path, or None if no data."""
    if snapshot_id is None:
        snapshot_id = latest_snapshot_id()
    if snapshot_id is None:
        return None

    with db() as conn:
        snap = conn.execute(
            "SELECT * FROM snapshots WHERE id = ?", (snapshot_id,)
        ).fetchone()
        if snap is None or snap["fetch_status"] != "success":
            return None

        # Resolve property info
        if prop is None:
            prop_id = snap["property_id"] if "property_id" in snap.keys() else None
            if prop_id:
                prop = get_property_by_id(prop_id)
        prop_name = prop["name"] if prop else PROPERTY_NAME
        prop_url = prop["url"] if prop else PROPERTY_URL

        units = [dict(r) for r in conn.execute(
            "SELECT * FROM units WHERE snapshot_id = ? "
            "ORDER BY beds, floorplan_name, unit_code",
            (snapshot_id,),
        ).fetchall()]

        floorplans = [dict(r) for r in conn.execute(
            "SELECT * FROM floorplans WHERE snapshot_id = ? "
            "ORDER BY beds, name",
            (snapshot_id,),
        ).fetchall()]

    if not units:
        return None

    snap = dict(snap)
    snap_date = datetime.fromisoformat(
        snap["fetched_at"].replace("Z", "+00:00")
    )
    snap_date_str = snap_date.strftime("%b %d, %Y")
    unit_count = snap["unit_count"]

    wb = Workbook()

    # -- Sheet 1: Availability Matrix --
    ws1 = wb.active
    ws1.title = "Availability Matrix"

    ws1.merge_cells("A1:G1")
    c = ws1["A1"]
    c.value = f"{prop_name} — Availability Matrix"
    c.font = _TITLE_FONT

    domain = prop_url.split("//")[-1].split("/")[0].replace("www.", "")
    ws1.merge_cells("A2:F2")
    c = ws1["A2"]
    c.value = f"Source: {domain}/floorplans  |  Snapshot: {snap_date_str}  |  Total available units: {unit_count}"
    c.font = _SUB_FONT

    headers = ["Unit Type", "Floorplan / Tier", "Unit Number",
               "Beds / Baths", "Sq Ft", "Rent", "Available Date"]
    _apply_header_row(ws1, 4, headers)

    by_bed: dict[int, list[dict]] = defaultdict(list)
    for u in units:
        by_bed[u["beds"]].append(u)

    row = 5
    for beds in sorted(by_bed.keys()):
        bed_units = by_bed[beds]
        label = _bed_label(beds)
        count = len(bed_units)

        ws1.merge_cells(f"A{row}:G{row}")
        c = ws1.cell(row=row, column=1, value=f"{label}  ({count} available)")
        c.font = _SECTION_FONT
        row += 1

        for u in bed_units:
            ws1.cell(row=row, column=1, value=label)
            ws1.cell(row=row, column=2, value=u["floorplan_name"])
            ws1.cell(row=row, column=3, value=u["unit_code"])
            ws1.cell(row=row, column=4, value=f"{u['beds']} / {int(u['baths'])}")
            c = ws1.cell(row=row, column=5, value=int(u["sqft"]))
            c.number_format = _NUM_FMT
            c = ws1.cell(row=row, column=6, value=int(u["min_rent"]))
            c.number_format = _MONEY_FMT
            try:
                dt = datetime.fromisoformat(u["available_date"].replace("Z", "+00:00"))
                c = ws1.cell(row=row, column=7, value=dt.replace(tzinfo=None))
                c.number_format = _DATE_FMT
            except (ValueError, AttributeError):
                ws1.cell(row=row, column=7, value=u.get("available_date", ""))
            row += 1

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 16

    # -- Sheet 2: Tier Summary --
    ws2 = wb.create_sheet("Tier Summary")

    ws2.merge_cells("A1:H1")
    c = ws2["A1"]
    c.value = "Floor Plan Tier Summary"
    c.font = _TITLE_FONT

    ws2.merge_cells("A2:H2")
    c = ws2["A2"]
    c.value = (f"{len(floorplans)} floor plan types  |  "
               f"{unit_count} available units  |  Snapshot: {snap_date_str}")
    c.font = _SUB_FONT

    headers2 = ["Tier / Floorplan", "Type", "Beds", "Baths",
                "Sq Ft", "Rent Range", "# Available", "Earliest Move-In"]
    _apply_header_row(ws2, 4, headers2)

    row = 5
    for fp in floorplans:
        label = _bed_label(fp["beds"])
        ws2.cell(row=row, column=1, value=fp["name"])
        ws2.cell(row=row, column=2, value=label)
        ws2.cell(row=row, column=3, value=fp["beds"])
        ws2.cell(row=row, column=4, value=int(fp["baths"]))
        c = ws2.cell(row=row, column=5, value=int(fp["min_sqft"]))
        c.number_format = _NUM_FMT

        mn, mx = int(fp["min_rent"]), int(fp["max_rent"])
        rent_str = f"${mn:,}" if mn == mx else f"${mn:,} – ${mx:,}"
        ws2.cell(row=row, column=6, value=rent_str)

        ws2.cell(row=row, column=7, value=fp["available_count"])

        try:
            dt = datetime.fromisoformat(fp["available_date"].replace("Z", "+00:00"))
            c = ws2.cell(row=row, column=8, value=dt.replace(tzinfo=None))
            c.number_format = _DATE_FMT
        except (ValueError, AttributeError, TypeError):
            ws2.cell(row=row, column=8, value=fp.get("available_date", ""))
        row += 1

    total_row = row
    ws2.cell(row=total_row, column=1, value="TOTAL")
    ws2.cell(row=total_row, column=7,
             value=f"=SUM(G5:G{total_row - 1})")

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 22
    ws2.column_dimensions["G"].width = 14
    ws2.column_dimensions["H"].width = 18

    # -- Sheet 3: Raw Data --
    ws3 = wb.create_sheet("Raw Data")

    raw_headers = ["Unit Code", "Floorplan", "Bed Type", "Beds", "Baths",
                   "Sq Ft", "Min Rent", "Max Rent", "Available Date", "Rent / Sq Ft"]
    _apply_header_row(ws3, 1, raw_headers)

    for i, u in enumerate(units, 2):
        ws3.cell(row=i, column=1, value=u["unit_code"])
        ws3.cell(row=i, column=2, value=u["floorplan_name"])
        ws3.cell(row=i, column=3, value=_bed_label(u["beds"]))
        ws3.cell(row=i, column=4, value=u["beds"])
        ws3.cell(row=i, column=5, value=int(u["baths"]))
        c = ws3.cell(row=i, column=6, value=int(u["sqft"]))
        c.number_format = _NUM_FMT
        c = ws3.cell(row=i, column=7, value=int(u["min_rent"]))
        c.number_format = _MONEY_FMT
        c = ws3.cell(row=i, column=8, value=int(u["max_rent"]))
        c.number_format = _MONEY_FMT
        try:
            dt = datetime.fromisoformat(u["available_date"].replace("Z", "+00:00"))
            c = ws3.cell(row=i, column=9, value=dt.replace(tzinfo=None))
            c.number_format = _DATE_FMT
        except (ValueError, AttributeError):
            ws3.cell(row=i, column=9, value=u.get("available_date", ""))
        c = ws3.cell(row=i, column=10, value=f"=G{i}/F{i}")
        c.number_format = _PSF_FMT

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 7
    ws3.column_dimensions["E"].width = 7
    ws3.column_dimensions["F"].width = 9
    ws3.column_dimensions["G"].width = 11
    ws3.column_dimensions["H"].width = 11
    ws3.column_dimensions["I"].width = 16
    ws3.column_dimensions["J"].width = 13

    # -- Save --
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    slug = _slugify(prop_name)
    out_path = EXPORTS_DIR / f"{slug}_Availability.xlsx"
    wb.save(str(out_path))
    return str(out_path)


# =============================================================================
# CONSOLIDATED COMP REPORT (all properties in one workbook)
# =============================================================================

def comp_report_path() -> Path:
    """Path for the consolidated comp report workbook."""
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return EXPORTS_DIR / "Rent_Comp_Report.xlsx"


def export_comp_report(snapshot_ids: dict[int, int | None] | None = None) -> str | None:
    """Build a multi-property comp report workbook.

    Sheets:
      1. Comp Overview    — cross-property comparison (rent, sqft, PSF, exposure, fees)
      2. Rents by Type    — per-bed-type breakdown across properties
      3. Concessions      — concession text and NER for each property
      4. Fees             — fee comparison table
      5–N. Per-property   — one sheet per property with unit-level detail

    Returns output path, or None if no data.
    """
    from src.config import get_active_properties
    from src.ner import calculate_ner

    properties = get_active_properties()
    if not properties:
        return None

    now = datetime.now(timezone.utc)
    date_str = now.strftime("%b %d, %Y")

    # Gather data for each property
    prop_data = []
    with db() as conn:
        for prop in properties:
            pid = prop["id"]

            # Use provided snapshot_id or find latest
            if snapshot_ids and pid in snapshot_ids and snapshot_ids[pid]:
                sid = snapshot_ids[pid]
            else:
                row = conn.execute(
                    "SELECT MAX(id) AS sid FROM snapshots "
                    "WHERE property_id = ? AND fetch_status = 'success'",
                    (pid,),
                ).fetchone()
                sid = row["sid"] if row else None

            if not sid:
                prop_data.append({"prop": prop, "snap": None, "units": [], "floorplans": []})
                continue

            snap = dict(conn.execute(
                "SELECT * FROM snapshots WHERE id = ?", (sid,)
            ).fetchone())

            units = [dict(r) for r in conn.execute(
                "SELECT * FROM units WHERE snapshot_id = ? "
                "ORDER BY beds, floorplan_name, unit_code",
                (sid,),
            ).fetchall()]

            floorplans = [dict(r) for r in conn.execute(
                "SELECT * FROM floorplans WHERE snapshot_id = ? "
                "ORDER BY beds, name",
                (sid,),
            ).fetchall()]

            prop_data.append({
                "prop": prop, "snap": snap, "units": units, "floorplans": floorplans,
            })

    if not any(d["units"] for d in prop_data):
        return None

    wb = Workbook()

    # -- Sheet 1: Comp Overview --
    _build_comp_overview(wb, prop_data, date_str)

    # -- Sheet 2: Rents by Unit Type --
    _build_rents_by_type_sheet(wb, prop_data, date_str)

    # -- Sheet 3: Concessions --
    _build_concessions_sheet(wb, prop_data, date_str, calculate_ner)

    # -- Sheet 4: Fees --
    _build_fees_sheet(wb, prop_data, date_str)

    # -- Sheet 5: Pricing Recommendations --
    _build_pricing_sheet(wb, date_str)

    # -- Per-property detail sheets --
    for d in prop_data:
        if d["units"]:
            _build_property_sheet(wb, d)

    out = comp_report_path()
    wb.save(str(out))
    return str(out)


def _build_comp_overview(wb: Workbook, prop_data: list[dict], date_str: str) -> None:
    """Sheet 1: Cross-property comparison overview."""
    ws = wb.active
    ws.title = "Comp Overview"

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "Rent Comp Report — Overview"
    c.font = _TITLE_FONT

    subject = next((d for d in prop_data if d["prop"].get("is_subject")), None)
    subject_name = subject["prop"]["name"] if subject else "Comp Set"
    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = f"Subject: {subject_name}  |  {len(prop_data)} properties  |  {date_str}"
    c.font = _SUB_FONT

    headers = [
        "Property", "Mgmt", "Year Built", "Total Units",
        "Available", "Exposure %", "Avg Rent", "Avg SqFt",
        "Rent/SqFt", "Concessions",
    ]
    _apply_header_row(ws, 4, headers)

    row = 5
    comp_rents = []
    comp_sqfts = []
    comp_psfs = []
    comp_exposures = []

    for d in prop_data:
        p = d["prop"]
        units = d["units"]
        is_subj = p.get("is_subject")

        total = p.get("unit_count_total") or 0
        avail = len(units)
        exposure = avail / total if total else 0

        rents = [u["min_rent"] for u in units if u["min_rent"]]
        sqfts = [u["sqft"] for u in units if u["sqft"]]
        avg_rent = sum(rents) / len(rents) if rents else 0
        avg_sqft = sum(sqfts) / len(sqfts) if sqfts else 0
        psf = avg_rent / avg_sqft if avg_sqft else 0

        # Concession text summary
        conc_texts = list({(u.get("concession_text") or "") for u in units
                          if (u.get("concession_text") or "").strip()})
        conc_summary = conc_texts[0][:60] if conc_texts else "—"

        ws.cell(row=row, column=1, value=p["name"])
        ws.cell(row=row, column=2, value=p.get("management_company", ""))
        ws.cell(row=row, column=3, value=p.get("year_built"))
        ws.cell(row=row, column=4, value=total)
        ws.cell(row=row, column=5, value=avail)
        c = ws.cell(row=row, column=6, value=exposure)
        c.number_format = _PCT_FMT
        c = ws.cell(row=row, column=7, value=round(avg_rent) if avg_rent else None)
        c.number_format = _MONEY_FMT
        c = ws.cell(row=row, column=8, value=round(avg_sqft) if avg_sqft else None)
        c.number_format = _NUM_FMT
        c = ws.cell(row=row, column=9, value=round(psf, 2) if psf else None)
        c.number_format = _PSF_FMT
        ws.cell(row=row, column=10, value=conc_summary)

        # Highlight subject row
        if is_subj:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = _SUBJECT_FILL
                ws.cell(row=row, column=col).font = _SUBJECT_FONT
        else:
            if avg_rent:
                comp_rents.append(avg_rent)
            if avg_sqft:
                comp_sqfts.append(avg_sqft)
            if psf:
                comp_psfs.append(psf)
            if total:
                comp_exposures.append(exposure)

        for col in range(1, 11):
            ws.cell(row=row, column=col).border = _THIN_BORDER

        row += 1

    # Comp average row
    if comp_rents:
        ws.cell(row=row, column=1, value="COMP AVERAGE")
        ws.cell(row=row, column=1).font = _AVG_FONT
        c = ws.cell(row=row, column=6,
                     value=sum(comp_exposures) / len(comp_exposures) if comp_exposures else None)
        c.number_format = _PCT_FMT
        c = ws.cell(row=row, column=7,
                     value=round(sum(comp_rents) / len(comp_rents)))
        c.number_format = _MONEY_FMT
        c = ws.cell(row=row, column=8,
                     value=round(sum(comp_sqfts) / len(comp_sqfts)) if comp_sqfts else None)
        c.number_format = _NUM_FMT
        c = ws.cell(row=row, column=9,
                     value=round(sum(comp_psfs) / len(comp_psfs), 2) if comp_psfs else None)
        c.number_format = _PSF_FMT
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = _AVG_FILL
            ws.cell(row=row, column=col).font = _AVG_FONT

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 11
    ws.column_dimensions["J"].width = 40
    ws.freeze_panes = "B5"


def _build_rents_by_type_sheet(wb: Workbook, prop_data: list[dict], date_str: str) -> None:
    """Sheet 2: Per-bed-type breakdown across all properties."""
    ws = wb.create_sheet("Rents by Type")

    ws.merge_cells("A1:L1")
    ws["A1"].value = "Rents by Unit Type"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:L2")
    ws["A2"].value = f"Snapshot: {date_str}"
    ws["A2"].font = _SUB_FONT

    for beds in [0, 1, 2, 3]:
        label = _bed_label(beds)
        # Find starting row for this bed type section
        start_row = ws.max_row + 2

        ws.merge_cells(f"A{start_row}:H{start_row}")
        ws.cell(row=start_row, column=1, value=label)
        ws.cell(row=start_row, column=1).font = _SECTION_FONT

        headers = ["Property", "# Available", "Min Rent", "Avg Rent",
                   "Max Rent", "Avg SqFt", "Rent/SqFt", "Subject"]
        _apply_header_row(ws, start_row + 1, headers)

        row = start_row + 2
        comp_rents = []
        has_data = False

        for d in prop_data:
            p = d["prop"]
            bed_units = [u for u in d["units"] if u["beds"] == beds]
            if not bed_units:
                continue
            has_data = True
            is_subj = p.get("is_subject")

            rents = [u["min_rent"] for u in bed_units if u["min_rent"]]
            sqfts = [u["sqft"] for u in bed_units if u["sqft"]]
            avg_rent = sum(rents) / len(rents) if rents else 0
            avg_sqft = sum(sqfts) / len(sqfts) if sqfts else 0
            psf = avg_rent / avg_sqft if avg_sqft else 0

            ws.cell(row=row, column=1, value=p["name"])
            ws.cell(row=row, column=2, value=len(bed_units))
            c = ws.cell(row=row, column=3, value=min(rents) if rents else None)
            c.number_format = _MONEY_FMT
            c = ws.cell(row=row, column=4, value=round(avg_rent) if avg_rent else None)
            c.number_format = _MONEY_FMT
            c = ws.cell(row=row, column=5, value=max(rents) if rents else None)
            c.number_format = _MONEY_FMT
            c = ws.cell(row=row, column=6, value=round(avg_sqft) if avg_sqft else None)
            c.number_format = _NUM_FMT
            c = ws.cell(row=row, column=7, value=round(psf, 2) if psf else None)
            c.number_format = _PSF_FMT
            ws.cell(row=row, column=8, value="Yes" if is_subj else "")

            if is_subj:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = _SUBJECT_FILL
            elif avg_rent:
                comp_rents.append(avg_rent)

            row += 1

        if not has_data:
            ws.cell(row=row, column=1, value="No listings")
            ws.cell(row=row, column=1).font = Font(italic=True, color="FF999999")
            row += 1

    ws.column_dimensions["A"].width = 28
    for col_letter in "BCDEFGH":
        ws.column_dimensions[col_letter].width = 13
    ws.freeze_panes = "B1"


def _build_concessions_sheet(wb: Workbook, prop_data: list[dict],
                              date_str: str, calculate_ner) -> None:
    """Sheet 3: Concessions and NER for each property."""
    ws = wb.create_sheet("Concessions")

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Concessions & Net Effective Rent"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Snapshot: {date_str}"
    ws["A2"].font = _SUB_FONT

    headers = ["Property", "Concession Text", "Avg Gross Rent",
               "Avg NER", "Concession %", "Free Months Equiv", "Subject"]
    _apply_header_row(ws, 4, headers)

    row = 5
    for d in prop_data:
        p = d["prop"]
        units = d["units"]
        is_subj = p.get("is_subject")

        if not units:
            ws.cell(row=row, column=1, value=p["name"])
            ws.cell(row=row, column=2, value="No data")
            ws.cell(row=row, column=2).font = Font(italic=True, color="FF999999")
            row += 1
            continue

        # Get unique concession texts
        conc_texts = list({(u.get("concession_text") or "") for u in units
                          if (u.get("concession_text") or "").strip()})
        conc_display = "; ".join(t[:80] for t in conc_texts) if conc_texts else "None"

        # Calculate average NER
        ner_results = []
        for u in units:
            if u["min_rent"]:
                nr = calculate_ner(u["min_rent"], u.get("concession_text", ""))
                ner_results.append(nr)

        if ner_results:
            avg_gross = sum(nr.gross_rent for nr in ner_results) / len(ner_results)
            avg_ner = sum(nr.ner for nr in ner_results) / len(ner_results)
            avg_pct = sum(nr.concession_pct for nr in ner_results) / len(ner_results)
            avg_fm = sum(nr.free_months for nr in ner_results) / len(ner_results)
        else:
            avg_gross = avg_ner = avg_pct = avg_fm = 0

        ws.cell(row=row, column=1, value=p["name"])
        ws.cell(row=row, column=2, value=conc_display)
        c = ws.cell(row=row, column=3, value=round(avg_gross) if avg_gross else None)
        c.number_format = _MONEY_FMT
        c = ws.cell(row=row, column=4, value=round(avg_ner) if avg_ner else None)
        c.number_format = _MONEY_FMT
        c = ws.cell(row=row, column=5, value=avg_pct if avg_pct else None)
        c.number_format = _PCT_FMT
        ws.cell(row=row, column=6, value=round(avg_fm, 1) if avg_fm else None)
        ws.cell(row=row, column=7, value="Yes" if is_subj else "")

        if is_subj:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = _SUBJECT_FILL

        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 10
    ws.freeze_panes = "B5"


def _build_fees_sheet(wb: Workbook, prop_data: list[dict], date_str: str) -> None:
    """Sheet 4: Fee comparison across properties."""
    ws = wb.create_sheet("Fees")

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Fee Comparison"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Source: HelloData Report #53 & property websites  |  {date_str}"
    ws["A2"].font = _SUB_FONT

    headers = ["Property", "Application Fee", "Admin Fee",
               "Monthly Pet Rent", "Parking", "Subject"]
    _apply_header_row(ws, 4, headers)

    row = 5
    for d in prop_data:
        p = d["prop"]
        is_subj = p.get("is_subject")
        fees = FEES.get(p["name"], {})

        app_fee = admin_fee = pet_rent = parking = "—"

        if fees.get("application"):
            for item in fees["application"]:
                if "application" in item["item"].lower():
                    app_fee = item["cost"]
                elif "admin" in item["item"].lower():
                    admin_fee = item["cost"]

        if fees.get("pets"):
            for item in fees["pets"]:
                if "monthly" in item["item"].lower() or "rent" in item["item"].lower():
                    pet_rent = item["cost"]

        if fees.get("parking"):
            costs = [item["cost"] for item in fees["parking"]]
            parking = costs[0] if costs else "—"

        ws.cell(row=row, column=1, value=p["name"])
        ws.cell(row=row, column=2, value=app_fee)
        ws.cell(row=row, column=3, value=admin_fee)
        ws.cell(row=row, column=4, value=pet_rent)
        ws.cell(row=row, column=5, value=parking)
        ws.cell(row=row, column=6, value="Yes" if is_subj else "")

        if is_subj:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = _SUBJECT_FILL

        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 10
    ws.freeze_panes = "B5"


_OVER_FILL = PatternFill(start_color="FFFFF0E0", end_color="FFFFF0E0", fill_type="solid")
_UNDER_FILL = PatternFill(start_color="FFE0FFE0", end_color="FFE0FFE0", fill_type="solid")
_WELL_ABOVE_FILL = PatternFill(start_color="FFFFE0E0", end_color="FFFFE0E0", fill_type="solid")

_SIGNAL_FILLS = {
    "overpriced": _OVER_FILL,
    "well_above": _WELL_ABOVE_FILL,
    "underpriced": _UNDER_FILL,
}


def _build_pricing_sheet(wb: Workbook, date_str: str) -> None:
    """Sheet 5: Pricing recommendations for the subject property."""
    from src.pricing import generate_recommendations, signal_label

    report = generate_recommendations()
    if report is None:
        return

    ws = wb.create_sheet("Pricing Recommendations")

    ws.merge_cells("A1:M1")
    ws["A1"].value = f"Pricing Recommendations — {report.subject_name}"
    ws["A1"].font = _TITLE_FONT

    s = report.summary
    impact_sign = "+" if s["monthly_revenue_impact"] >= 0 else ""
    ws.merge_cells("A2:M2")
    ws["A2"].value = (
        f"Quality: {s['subject_quality_score']}/10 (comp avg {s['comp_avg_quality_score']})  |  "
        f"Exposure: {s['subject_exposure']}% (comp avg {s['comp_avg_exposure']}%)  |  "
        f"Revenue Impact: {impact_sign}${s['monthly_revenue_impact']:,}/mo  |  "
        f"{report.comp_count} comps  |  {date_str}"
    )
    ws["A2"].font = _SUB_FONT

    headers = [
        "Unit", "Floorplan", "Type", "SqFt", "Listed Rent",
        "Recommended", "Aggressive", "Conservative",
        "Delta ($)", "Delta (%)", "Signal",
        "Listed PSF", "Rec PSF",
    ]
    _apply_header_row(ws, 4, headers)

    bed_label = lambda b: "Studio" if b == 0 else f"{b}BR"

    for i, u in enumerate(report.units, 5):
        ws.cell(row=i, column=1, value=u.unit_code)
        ws.cell(row=i, column=2, value=u.floorplan_name)
        ws.cell(row=i, column=3, value=bed_label(u.beds))
        c = ws.cell(row=i, column=4, value=int(u.sqft))
        c.number_format = _NUM_FMT
        c = ws.cell(row=i, column=5, value=int(u.listed_rent))
        c.number_format = _MONEY_FMT
        c = ws.cell(row=i, column=6, value=int(u.recommended_rent))
        c.number_format = _MONEY_FMT
        c.font = Font(bold=True, size=11)
        c = ws.cell(row=i, column=7, value=int(u.aggressive_rent))
        c.number_format = _MONEY_FMT
        c = ws.cell(row=i, column=8, value=int(u.conservative_rent))
        c.number_format = _MONEY_FMT
        c = ws.cell(row=i, column=9, value=int(u.delta))
        c.number_format = _MONEY_FMT
        c = ws.cell(row=i, column=10, value=u.delta_pct)
        c.number_format = _PCT_FMT
        ws.cell(row=i, column=11, value=signal_label(u.signal))
        c = ws.cell(row=i, column=12, value=u.unit_psf)
        c.number_format = _PSF_FMT
        c = ws.cell(row=i, column=13, value=u.recommended_psf)
        c.number_format = _PSF_FMT

        # Color-code signal rows
        fill = _SIGNAL_FILLS.get(u.signal)
        if fill:
            for col in range(1, 14):
                ws.cell(row=i, column=col).fill = fill

    # Summary row
    summary_row = 5 + len(report.units) + 1
    ws.cell(row=summary_row, column=1, value="SUMMARY")
    ws.cell(row=summary_row, column=1).font = _AVG_FONT
    c = ws.cell(row=summary_row, column=5, value=s["avg_listed"])
    c.number_format = _MONEY_FMT
    c.font = _AVG_FONT
    c = ws.cell(row=summary_row, column=6, value=s["avg_recommended"])
    c.number_format = _MONEY_FMT
    c.font = _AVG_FONT
    c = ws.cell(row=summary_row, column=9, value=s["avg_delta"])
    c.number_format = _MONEY_FMT
    c.font = _AVG_FONT
    c = ws.cell(row=summary_row, column=10, value=s["avg_delta_pct"])
    c.number_format = _PCT_FMT
    c.font = _AVG_FONT
    ws.cell(row=summary_row, column=11,
            value=f"{s['overpriced_count']} over / {s['market_count']} market / {s['underpriced_count']} under")
    ws.cell(row=summary_row, column=11).font = _AVG_FONT

    for col in range(1, 14):
        ws.cell(row=summary_row, column=col).fill = _AVG_FILL

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 12
    ws.freeze_panes = "A5"


def _build_property_sheet(wb: Workbook, d: dict) -> None:
    """One detail sheet per property with unit-level data."""
    p = d["prop"]
    units = d["units"]
    snap = d["snap"]

    # Sheet name max 31 chars
    sheet_name = p["name"][:31]
    # Avoid duplicate sheet names
    existing = {ws.title for ws in wb.worksheets}
    if sheet_name in existing:
        sheet_name = sheet_name[:28] + "..."
    ws = wb.create_sheet(sheet_name)

    snap_date = ""
    if snap:
        try:
            dt = datetime.fromisoformat(snap["fetched_at"].replace("Z", "+00:00"))
            snap_date = dt.strftime("%b %d, %Y")
        except Exception:
            snap_date = snap["fetched_at"][:10]

    total = p.get("unit_count_total") or 0
    avail = len(units)
    exposure = f"{avail / total * 100:.1f}%" if total else "—"

    ws.merge_cells("A1:H1")
    ws["A1"].value = p["name"]
    ws["A1"].font = _TITLE_FONT

    ws.merge_cells("A2:H2")
    ws["A2"].value = (f"{p.get('address', '')}  |  "
                      f"{p.get('management_company', '')}  |  "
                      f"{total} total units  |  {avail} available  |  "
                      f"Exposure: {exposure}  |  {snap_date}")
    ws["A2"].font = _SUB_FONT

    headers = ["Unit Code", "Floorplan", "Beds", "Baths",
               "Sq Ft", "Rent", "Available Date", "Concession"]
    _apply_header_row(ws, 4, headers)

    for i, u in enumerate(units, 5):
        ws.cell(row=i, column=1, value=u["unit_code"])
        ws.cell(row=i, column=2, value=u["floorplan_name"])
        ws.cell(row=i, column=3, value=u["beds"])
        ws.cell(row=i, column=4, value=int(u["baths"]))
        c = ws.cell(row=i, column=5, value=int(u["sqft"]))
        c.number_format = _NUM_FMT
        c = ws.cell(row=i, column=6, value=int(u["min_rent"]))
        c.number_format = _MONEY_FMT
        try:
            dt = datetime.fromisoformat(u["available_date"].replace("Z", "+00:00"))
            c = ws.cell(row=i, column=7, value=dt.replace(tzinfo=None))
            c.number_format = _DATE_FMT
        except (ValueError, AttributeError):
            ws.cell(row=i, column=7, value=u.get("available_date", ""))
        ws.cell(row=i, column=8, value=u.get("concession_text", ""))

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 40
    ws.freeze_panes = "A5"
